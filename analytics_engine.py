"""
NSE Bhav Copy Momentum Analytics Engine
=========================================
OUTPUT: trend_output.xlsx  with FIVE SHEETS

  Sheet 1 — "Trend (True-False)"
    Company Name | Symbol | ISIN | Sector | CapCategory | Trending Days |
    12d Change% | 5d Change% |
    [today] ... [12th day]  ← True/False per day |
    ADX 14 | RSI 14 | EMA Signal

  Sheet 2 — "Trend (% Change)"
    Same columns but daily cells show % change instead of True/False.
    Positive % = green fill, Negative % = red fill.

  Sheet 3 — "Sector Momentum (12d)"
  Sheet 4 — "Sector Momentum (5d)"
  Sheet 5 — "Sector Momentum (Today)"

EMA SIGNAL LEGEND
─────────────────
  ██ DARK GREEN   "Golden Cross ✓"   — 50 EMA crossed ABOVE 200 EMA within
                                       the last 5 trading days. Strong buy.
  ██ LIGHT GREEN  "Above 200 EMA"    — 50 EMA already above 200 EMA (crossed
                                       > 5 days ago). Uptrend confirmed.
  ██ AMBER        "Approaching"      — 50 EMA is below 200 EMA but the gap is
                                       narrowing (50 EMA trending up toward
                                       200 EMA). Watch closely.
  ── (blank)      No signal          — 50 EMA below 200 EMA, gap widening.

SERIES FILTER: EQ and BE both included.
  EQ = regular equity settlement
  BE = Book Entry / Trade-to-Trade (delivery mandatory, no intraday)
  Both are genuine listed equity stocks valid for momentum analysis.

COLUMN DEFINITIONS
──────────────────
  12d Change%  = (Close_today − Close_13th_day_back) / Close_13th_day_back × 100
  5d Change%   = (Close_today − Close_6th_day_back)  / Close_6th_day_back  × 100
  Daily % cell = (Close_day  − Close_prev_day)       / Close_prev_day      × 100
  EMA 50       = Exponential Moving Average over 50 trading days
  EMA 200      = Exponential Moving Average over 200 trading days
  EMA Signal   = Classification based on EMA50 vs EMA200 relationship

SETUP
─────
  1. Set DATA_FOLDER       → folder with all YYYYMMDD_NSE.csv files
  2. Set NSE_MASTER_CSV    → path to EQUITY_L.csv  (nseindia.com)
  3. Set NSE_SECTOR_MASTER → path to nse_sector_master.csv
  4. Run:  python analytics_engine.py
  5. Output: YYYYMMDD_trend_output.xlsx
"""

import os
import re
import warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from datetime import date

warnings.filterwarnings("ignore")

today_str = date.today().strftime("%Y%m%d")

# ─────────────────────────────────────────────────────────────────────
#  CONFIGURATION  — edit these paths before running
# ─────────────────────────────────────────────────────────────────────
DATA_FOLDER       = r"C:\Users\user-pc\Downloads\200_Bhav"
OUTPUT_FILE       = f"{today_str}_trend_output.xlsx"
NSE_MASTER_CSV    = r"C:\Users\user-pc\Downloads\EQUITY_L.csv"
NSE_SECTOR_MASTER = r"C:\Users\user-pc\Downloads\nse_sector_master.csv"

LOOKBACK_DAYS    = 12
RSI_PERIOD       = 14
ADX_PERIOD       = 14
MIN_ROWS_ADX     = ADX_PERIOD * 2
MIN_TOTAL_TRADES = 0
W52_DAYS         = 252

# EMA parameters
EMA_FAST         = 50    # fast EMA period
EMA_SLOW         = 200   # slow EMA period
EMA_CROSS_WINDOW = 5     # days to look back for "just crossed"
EMA_MIN_ROWS     = EMA_SLOW + 10  # need 210 rows for reliable EMA 200

# Series to include — EQ (regular) + BE (book entry / trade-to-trade)
ALLOWED_SERIES   = {"EQ", "BE"}


# ─────────────────────────────────────────────────────────────────────
#  STEP 1 — LOAD ALL BHAV FILES
# ─────────────────────────────────────────────────────────────────────
def load_bhav_files(folder):
    pattern = re.compile(r"^(\d{8})_NSE\.csv$", re.IGNORECASE)
    frames  = []

    for fname in sorted(os.listdir(folder)):
        m = pattern.match(fname)
        if not m:
            continue
        trade_date = pd.to_datetime(m.group(1), format="%Y%m%d")
        fpath = os.path.join(folder, fname)
        try:
            df = pd.read_csv(fpath, dtype=str)
        except Exception as e:
            print(f"  [WARN] Cannot read {fname}: {e}")
            continue

        df.columns = df.columns.str.strip().str.upper()
        required = {"SYMBOL", "SERIES", "HIGH", "LOW", "CLOSE"}
        if not required.issubset(df.columns):
            print(f"  [WARN] {fname} missing cols — skipped")
            continue

        # ── SERIES FILTER: EQ + BE ────────────────────────────────
        total_before = len(df)
        df["SERIES"] = df["SERIES"].str.strip().str.upper()
        df = df[df["SERIES"].isin(ALLOWED_SERIES)].copy()
        print(f"  {fname}: {total_before} rows → {len(df)} EQ+BE rows")
        if df.empty:
            continue

        if MIN_TOTAL_TRADES > 0 and "TOTALTRADES" in df.columns:
            df["TOTALTRADES"] = pd.to_numeric(
                df["TOTALTRADES"], errors="coerce"
            ).fillna(0)
            df = df[df["TOTALTRADES"] >= MIN_TOTAL_TRADES].copy()
            if df.empty:
                continue

        df["DATE"] = trade_date
        for col in ["HIGH", "LOW", "CLOSE"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")

        keep = ["SYMBOL", "SERIES", "DATE", "HIGH", "LOW", "CLOSE"]
        if "ISIN" in df.columns:
            keep.append("ISIN")
        if "TOTTRDQTY" in df.columns:
            keep.append("TOTTRDQTY")
        if "TOTALTRADES" in df.columns:
            keep.append("TOTALTRADES")
        frames.append(df[keep])

    if not frames:
        raise RuntimeError(f"No YYYYMMDD_NSE.csv files found in '{folder}'")

    combined = pd.concat(frames, ignore_index=True)

    # Safety check
    leaked = combined[~combined["SERIES"].isin(ALLOWED_SERIES)]
    if not leaked.empty:
        raise AssertionError(
            f"Unexpected series leaked: {leaked['SERIES'].unique().tolist()}"
        )
    combined = combined.drop(columns=["SERIES"])

    # ISIN map — last known ISIN per symbol
    if "ISIN" in combined.columns:
        isin_map = (
            combined.sort_values("DATE")
            .dropna(subset=["ISIN"])
            .groupby("SYMBOL")["ISIN"].last()
        )
        combined = combined.drop(columns=["ISIN"])
    else:
        isin_map = pd.Series(dtype=str)

    all_dates = sorted(combined["DATE"].unique())
    today_ts  = pd.Timestamp(all_dates[-1])
    oldest_ts = pd.Timestamp(all_dates[0])

    print(f"\n{'─'*55}")
    print(f"  Files loaded       : {len(frames)}")
    print(f"  Trading dates      : {len(all_dates)}")
    print(f"  EQ+BE symbols      : {combined['SYMBOL'].nunique()}")
    print(f"  Oldest date        : {oldest_ts.strftime('%d %b %Y')}")
    print(f"  TODAY (newest)     : {today_ts.strftime('%d %b %Y')}")
    print(f"{'─'*55}\n")

    return combined, isin_map


# ─────────────────────────────────────────────────────────────────────
#  STEP 2 — BUILD PRICE PIVOTS
# ─────────────────────────────────────────────────────────────────────
def build_price_pivots(df):
    df = df.sort_values(["SYMBOL", "DATE"])
    kw = dict(aggfunc="last")
    close = df.pivot_table(index="SYMBOL", columns="DATE", values="CLOSE", **kw)
    high  = df.pivot_table(index="SYMBOL", columns="DATE", values="HIGH",  **kw)
    low   = df.pivot_table(index="SYMBOL", columns="DATE", values="LOW",   **kw)
    return close, high, low


# ─────────────────────────────────────────────────────────────────────
#  STEP 3 — MATRIX DATES
# ─────────────────────────────────────────────────────────────────────
def get_matrix_dates(close_pivot, lookback):
    all_dates = close_pivot.columns
    if len(all_dates) < lookback + 1:
        raise RuntimeError(
            f"Need ≥{lookback+1} trading days. Found {len(all_dates)}."
        )
    matrix_dates = all_dates[-lookback:]
    print(f"  12-day matrix  : {matrix_dates[0].strftime('%Y-%m-%d')} → "
          f"{matrix_dates[-1].strftime('%Y-%m-%d')} (TODAY)")
    print(f"  Prev-close ref : {all_dates[-(lookback+1)].strftime('%Y-%m-%d')}\n")
    return matrix_dates


# ─────────────────────────────────────────────────────────────────────
#  STEP 4 — TRENDING DAYS + TRUE/FALSE + DAILY % MATRIX
# ─────────────────────────────────────────────────────────────────────
def compute_matrices(close_pivot, matrix_dates):
    change_abs = close_pivot.diff(axis=1)
    pct_change = (change_abs / close_pivot.shift(1, axis=1)) * 100

    bool_matrix = (change_abs > 0)[matrix_dates].copy()
    pct_matrix  = pct_change[matrix_dates].copy().round(2)
    trending    = bool_matrix.sum(axis=1).astype(int)
    trending.name = "Trending Days"

    return trending, bool_matrix, pct_matrix


# ─────────────────────────────────────────────────────────────────────
#  STEP 5 — 12d AND 5d PERCENT CHANGE
# ─────────────────────────────────────────────────────────────────────
def compute_period_changes(close_pivot, matrix_dates):
    all_dates   = close_pivot.columns
    today_date  = matrix_dates[-1]
    today_idx   = list(all_dates).index(today_date)

    ref_5d_idx  = today_idx - 6
    ref_12d_idx = today_idx - 12

    close_today = close_pivot.iloc[:, today_idx]

    if ref_12d_idx >= 0:
        close_12d_ref = close_pivot.iloc[:, ref_12d_idx]
        chg_12d = ((close_today - close_12d_ref) / close_12d_ref * 100).round(2)
        print(f"  12d Change% : {all_dates[ref_12d_idx].strftime('%Y-%m-%d')} → "
              f"{today_date.strftime('%Y-%m-%d')}")
    else:
        chg_12d = pd.Series(np.nan, index=close_pivot.index)
        print("  12d Change% : insufficient history")
    chg_12d.name = "12d Change%"

    if ref_5d_idx >= 0:
        close_5d_ref = close_pivot.iloc[:, ref_5d_idx]
        chg_5d = ((close_today - close_5d_ref) / close_5d_ref * 100).round(2)
        print(f"  5d  Change% : {all_dates[ref_5d_idx].strftime('%Y-%m-%d')} → "
              f"{today_date.strftime('%Y-%m-%d')}\n")
    else:
        chg_5d = pd.Series(np.nan, index=close_pivot.index)
        print("  5d  Change% : insufficient history\n")
    chg_5d.name = "5d Change%"

    return chg_12d, chg_5d


# ─────────────────────────────────────────────────────────────────────
#  WILDER SMOOTHING — used by RSI and ADX
# ─────────────────────────────────────────────────────────────────────
def wilder_smooth(series, period):
    result = np.full(len(series), np.nan)
    if len(series) < period:
        return result
    result[period - 1] = np.nanmean(series[:period])
    alpha = 1.0 / period
    for i in range(period, len(series)):
        prev = result[i - 1]
        val  = series[i]
        result[i] = (
            prev * (1 - alpha) + val * alpha if not np.isnan(val) else prev
        )
    return result


# ─────────────────────────────────────────────────────────────────────
#  STEP 6 — RSI
# ─────────────────────────────────────────────────────────────────────
def compute_rsi_for_symbol(closes, period=14):
    if len(closes) < period + 1:
        return np.nan
    delta    = np.diff(closes.astype(float))
    avg_gain = wilder_smooth(np.where(delta > 0, delta,  0.0), period)
    avg_loss = wilder_smooth(np.where(delta < 0, -delta, 0.0), period)
    g, l = avg_gain[-1], avg_loss[-1]
    if np.isnan(g) or np.isnan(l):
        return np.nan
    return 100.0 if l == 0 else round(100 - (100 / (1 + g / l)), 2)


def compute_rsi(close_pivot, period=14):
    return pd.Series(
        {sym: compute_rsi_for_symbol(row.dropna().values, period)
         for sym, row in close_pivot.iterrows()},
        name="RSI 14",
    ).round(2)


# ─────────────────────────────────────────────────────────────────────
#  STEP 7 — ADX
# ─────────────────────────────────────────────────────────────────────
def compute_adx_for_symbol(h, l, c, period=14, min_rows=28):
    if len(c) < min_rows:
        return np.nan
    h, l, c = h.astype(float), l.astype(float), c.astype(float)
    tr  = np.maximum(
        h[1:] - l[1:],
        np.maximum(np.abs(h[1:] - c[:-1]), np.abs(l[1:] - c[:-1]))
    )
    up  = h[1:] - h[:-1]
    dn  = l[:-1] - l[1:]
    pdm = np.where((up > dn) & (up > 0), up, 0.0)
    ndm = np.where((dn > up) & (dn > 0), dn, 0.0)
    s_tr  = wilder_smooth(tr,  period)
    s_pdm = wilder_smooth(pdm, period)
    s_ndm = wilder_smooth(ndm, period)
    with np.errstate(divide="ignore", invalid="ignore"):
        pdi = 100 * np.where(s_tr > 0, s_pdm / s_tr, 0.0)
        ndi = 100 * np.where(s_tr > 0, s_ndm / s_tr, 0.0)
        s   = pdi + ndi
        dx  = 100 * np.where(s > 0, np.abs(pdi - ndi) / s, 0.0)
    adx = wilder_smooth(dx, period)
    return round(float(adx[-1]), 2) if not np.isnan(adx[-1]) else np.nan


def compute_adx(close_piv, high_piv, low_piv, period=14, min_rows=28):
    dates        = close_piv.columns
    high_aligned = high_piv.reindex(columns=dates)
    low_aligned  = low_piv.reindex(columns=dates)
    result = {}
    for sym in close_piv.index:
        c     = close_piv.loc[sym]
        h     = (high_aligned.loc[sym]
                 if sym in high_aligned.index else pd.Series(dtype=float))
        l     = (low_aligned.loc[sym]
                 if sym in low_aligned.index  else pd.Series(dtype=float))
        valid = c.notna() & h.notna() & l.notna()
        result[sym] = compute_adx_for_symbol(
            h[valid].values, l[valid].values, c[valid].values, period, min_rows
        )
    return pd.Series(result, name="ADX 14").round(2)


# ─────────────────────────────────────────────────────────────────────
#  STEP 8 — EMA 50 / EMA 200 + GOLDEN CROSS SIGNAL
# ─────────────────────────────────────────────────────────────────────
# Signal values stored in "EMA Signal" column:
#   "golden_cross"  — 50 EMA just crossed above 200 EMA (within last N days)
#   "above_200"     — 50 EMA already above 200 EMA (confirmed uptrend)
#   "approaching"   — 50 EMA below 200 EMA but gap is narrowing
#   ""              — no signal (50 EMA below 200, gap widening or flat)
#
# EMA 50 and EMA 200 current values also stored as separate columns
# for sorting/filtering in the screener.
# ─────────────────────────────────────────────────────────────────────

def compute_ema_signals(
    close_piv: pd.DataFrame,
    fast:        int = EMA_FAST,
    slow:        int = EMA_SLOW,
    cross_window: int = EMA_CROSS_WINDOW,
    min_rows:    int = EMA_MIN_ROWS,
) -> pd.DataFrame:
    """
    Compute EMA(50), EMA(200), and a golden-cross signal for every symbol.

    Parameters
    ----------
    close_piv    : DataFrame  shape (symbols × dates), oldest→newest columns
    fast         : fast EMA period  (default 50)
    slow         : slow EMA period  (default 200)
    cross_window : look-back window in trading days to detect a fresh cross
    min_rows     : minimum non-null history required for valid EMA 200

    Returns
    -------
    DataFrame with index = symbol, columns:
        EMA 50      — latest EMA(50) value
        EMA 200     — latest EMA(200) value
        EMA Signal  — "golden_cross" | "above_200" | "approaching" | ""

    Implementation
    ──────────────
    Uses pandas ewm(span=N, adjust=False) — same as TradingView EMA.

    Golden cross detection:
        We look at the last (cross_window + 1) values of both EMA series.
        A cross happened if, at some point in that window, EMA_fast
        transitioned from below EMA_slow to above EMA_slow.

    Approaching detection:
        EMA_fast < EMA_slow  BUT  the gap (EMA_slow - EMA_fast) has been
        shrinking over the last cross_window days — i.e., the 50 EMA is
        trending up toward the 200 EMA.
    """
    # Transpose: rows = dates, columns = symbols — ewm runs down each column
    prices = close_piv.T

    ema_fast_df = prices.ewm(span=fast, adjust=False).mean()
    ema_slow_df = prices.ewm(span=slow, adjust=False).mean()

    # Null out symbols with insufficient history
    valid_counts = close_piv.notna().sum(axis=1)

    records = {}
    for sym in close_piv.index:
        if valid_counts[sym] < min_rows:
            records[sym] = {
                "EMA 50":     np.nan,
                "EMA 200":    np.nan,
                "EMA Signal": "",
            }
            continue

        ef = ema_fast_df[sym].dropna().values   # fast EMA time series
        es = ema_slow_df[sym].dropna().values   # slow EMA time series

        if len(ef) < cross_window + 1 or len(es) < cross_window + 1:
            records[sym] = {
                "EMA 50":     np.nan,
                "EMA 200":    np.nan,
                "EMA Signal": "",
            }
            continue

        ema50_now  = round(float(ef[-1]), 4)
        ema200_now = round(float(es[-1]), 4)

        # ── Detect golden cross in the last cross_window days ─────────
        # Look at the window of [-(cross_window+1) : ] for both series
        ef_window = ef[-(cross_window + 1):]
        es_window = es[-(cross_window + 1):]
        # diff[i] = fast[i] - slow[i] at each point in window
        diff_window = ef_window - es_window

        # A cross occurred if sign changed from negative to positive
        # anywhere in the window (i.e., was below, now above)
        was_below = diff_window[:-1] < 0   # each day before last
        now_above = diff_window[1:] > 0    # each following day
        fresh_cross = bool(np.any(was_below & now_above))

        # ── Classify ──────────────────────────────────────────────────
        if fresh_cross:
            signal = "golden_cross"

        elif ema50_now > ema200_now:
            # Already above — confirmed uptrend
            signal = "above_200"

        else:
            # 50 EMA below 200 EMA — check if gap is narrowing
            # Gap at each of the last cross_window days
            gaps = ef[-(cross_window + 1):] - es[-(cross_window + 1):]
            # Narrowing = gap is becoming less negative (slope > 0)
            gap_slope = np.polyfit(range(len(gaps)), gaps, 1)[0]
            signal = "approaching" if gap_slope > 0 else ""

        records[sym] = {
            "EMA 50":     ema50_now,
            "EMA 200":    ema200_now,
            "EMA Signal": signal,
        }

    result = pd.DataFrame.from_dict(records, orient="index")
    result.index.name = "Symbol"
    return result


# ─────────────────────────────────────────────────────────────────────
#  52-WEEK HIGH
# ─────────────────────────────────────────────────────────────────────
def compute_52w_high(close_pivot, w52_days=252):
    recent = close_pivot.iloc[:, -w52_days:]
    high_52w      = recent.max(axis=1)
    high_52w.name = "52W High"
    close_today   = close_pivot.iloc[:, -1]
    pct_from_high = ((close_today - high_52w) / high_52w * 100).round(2)
    pct_from_high.name = "52W High %"
    near_high     = (pct_from_high >= -5)
    near_high.name = "Near 52W High"
    rank_52w      = (
        close_today.rank(pct=True) * 100
    ).round(1)
    rank_52w.name = "52W Rank"
    return high_52w, pct_from_high, near_high, rank_52w


# ─────────────────────────────────────────────────────────────────────
#  RELATIVE STRENGTH SCORE
# ─────────────────────────────────────────────────────────────────────
def compute_rs_score(close_pivot):
    """Percentile rank of 12-month return vs all other symbols."""
    if close_pivot.shape[1] < 2:
        return pd.Series(np.nan, index=close_pivot.index, name="RS Score")
    ret = (close_pivot.iloc[:, -1] / close_pivot.iloc[:, 0] - 1) * 100
    rs  = (ret.rank(pct=True) * 100).round(1)
    rs.name = "RS Score"
    return rs


# ─────────────────────────────────────────────────────────────────────
#  MOMENTUM SCORE (composite 0–100)
# ─────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────
#  WEIGHTED RPI  (Relative Price Index — IBD/Mansfield style)
# ─────────────────────────────────────────────────────────────────────
def compute_weighted_rpi(close_pivot):
    """
    Weighted Relative Price Index — ranks each stock's composite return
    vs all other stocks.  More weight on recent periods.

    Formula (IBD-style):
      composite = 1m_return×0.40 + 3m_return×0.20 + 6m_return×0.20 + 12m_return×0.20

    Returns percentile rank 1–99 (99 = strongest relative performer).
    Requires at least 21 trading days (≈1 month) of history.
    """
    if close_pivot.shape[1] < 21:
        return pd.Series(np.nan, index=close_pivot.index, name="Weighted RPI")

    today   = close_pivot.iloc[:, -1]
    n_dates = close_pivot.shape[1]

    def _ret(n_days):
        """Return vs n_days ago, or NaN if insufficient history."""
        if n_dates <= n_days:
            return pd.Series(np.nan, index=close_pivot.index)
        ref = close_pivot.iloc[:, -(n_days + 1)]
        return ((today - ref) / ref * 100).replace([np.inf, -np.inf], np.nan)

    # Approximate trading days per period
    r1m  = _ret(21)    # ~1 month
    r3m  = _ret(63)    # ~3 months
    r6m  = _ret(126)   # ~6 months
    r12m = _ret(252)   # ~12 months

    # Weighted composite — fill missing longer periods with 1m return
    composite = (
        r1m.fillna(0)  * 0.40 +
        r3m.fillna(r1m.fillna(0))  * 0.20 +
        r6m.fillna(r3m.fillna(r1m.fillna(0)))  * 0.20 +
        r12m.fillna(r6m.fillna(r3m.fillna(r1m.fillna(0))))  * 0.20
    )

    # Percentile rank 1–99
    rpi = (composite.rank(pct=True) * 98 + 1).clip(1, 99).round(1)
    rpi.name = "Weighted RPI"
    return rpi



# ─────────────────────────────────────────────────────────────────────
#  MULTI-PERIOD RPI  (2W / 3M / 6M / 6M-SMA2W)
# ─────────────────────────────────────────────────────────────────────
def compute_rpi_periods(close_pivot: pd.DataFrame) -> pd.DataFrame:
    """
    Compute Relative Price Index for 4 periods, each percentile-ranked 1-99.

    Periods (trading days):
      2W       = 10  days
      3M       = 63  days
      6M       = 126 days
      6M-SMA2W = simple average of today's 6M RPI and the 6M RPI from 10 days ago

    Returns DataFrame indexed by Symbol:
      RPI 2W | RPI 3M | RPI 6M | RPI 6M SMA2W
    """
    if close_pivot.shape[1] < 11:
        empty = pd.Series(np.nan, index=close_pivot.index)
        return pd.DataFrame({
            "RPI 2W": empty, "RPI 3M": empty,
            "RPI 6M": empty, "RPI 6M SMA2W": empty,
        })

    today  = close_pivot.iloc[:, -1]
    n_cols = close_pivot.shape[1]

    def _rpi(n_days):
        if n_cols <= n_days:
            return pd.Series(np.nan, index=close_pivot.index)
        ref = close_pivot.iloc[:, -(n_days + 1)]
        ret = ((today - ref) / ref * 100).replace([np.inf, -np.inf], np.nan)
        return (ret.rank(pct=True) * 98 + 1).clip(1, 99).round(1)

    rpi_2w = _rpi(10)
    rpi_3m = _rpi(63)
    rpi_6m = _rpi(126)

    # 6M SMA 2W — average of 6M RPI today vs 6M RPI from 10 days ago
    if n_cols > 137:
        piv_ago  = close_pivot.iloc[:, :-10]
        today_ago = piv_ago.iloc[:, -1]
        if piv_ago.shape[1] > 126:
            ref_ago  = piv_ago.iloc[:, -(126 + 1)]
            ret_ago  = ((today_ago - ref_ago) / ref_ago * 100).replace([np.inf, -np.inf], np.nan)
            rpi_ago  = (ret_ago.rank(pct=True) * 98 + 1).clip(1, 99).round(1)
            rpi_sma  = ((rpi_6m + rpi_ago) / 2).round(1)
        else:
            rpi_sma = rpi_6m.copy()
    else:
        rpi_sma = rpi_6m.copy()

    df = pd.DataFrame({
        "RPI 2W":       rpi_2w,
        "RPI 3M":       rpi_3m,
        "RPI 6M":       rpi_6m,
        "RPI 6M SMA2W": rpi_sma,
    })
    df.index.name = "Symbol"
    return df


# ─────────────────────────────────────────────────────────────────────
#  EMA 9 AND EMA 21  (Super Strength indicators)
# ─────────────────────────────────────────────────────────────────────
def compute_ema_short(close_pivot: pd.DataFrame) -> pd.DataFrame:
    """
    Compute EMA(9) and EMA(21).  Uses ewm(adjust=False) = TradingView formula.
    Returns DataFrame: EMA 9 | EMA 21
    """
    prices   = close_pivot.T
    ema9_df  = prices.ewm(span=9,  adjust=False).mean()
    ema21_df = prices.ewm(span=21, adjust=False).mean()
    valid    = close_pivot.notna().sum(axis=1)

    df = pd.DataFrame({
        "EMA 9":  ema9_df.iloc[-1].round(4),
        "EMA 21": ema21_df.iloc[-1].round(4),
    })
    df.index.name = "Symbol"
    df.loc[valid < 9,  "EMA 9"]  = np.nan
    df.loc[valid < 21, "EMA 21"] = np.nan
    return df


# ─────────────────────────────────────────────────────────────────────
#  SHORT-PERIOD RSI  (1D and 1W)
# ─────────────────────────────────────────────────────────────────────
def compute_rsi_short(close_pivot: pd.DataFrame) -> pd.DataFrame:
    """
    RSI(1)  — 1-day momentum pulse (>50 = up day, <50 = down day)
    RSI(5)  — 1-week (5-day) momentum.  > 60 = bullish short-term.
    Returns DataFrame: RSI 1D | RSI 1W
    """
    rsi_1d = pd.Series(
        {s: compute_rsi_for_symbol(r.dropna().values, period=1)
         for s, r in close_pivot.iterrows()},
        name="RSI 1D",
    ).round(2)
    rsi_1w = pd.Series(
        {s: compute_rsi_for_symbol(r.dropna().values, period=5)
         for s, r in close_pivot.iterrows()},
        name="RSI 1W",
    ).round(2)
    df = pd.DataFrame({"RSI 1D": rsi_1d, "RSI 1W": rsi_1w})
    df.index.name = "Symbol"
    return df


def compute_momentum_score(trending, rsi, adx, rs_score, pct_from_high):
    """
    Weighted composite of five signals, each normalised to 0–1:
      Trending Days (30%) · RSI (20%) · ADX (20%) · RS Score (20%) · 52W% (10%)
    """
    td_norm  = trending / 12
    rsi_norm = (rsi.reindex(trending.index)    .fillna(50) / 100).clip(0, 1)
    adx_norm = (adx.reindex(trending.index)    .fillna(0)  / 60).clip(0, 1)
    rs_norm  = (rs_score.reindex(trending.index).fillna(50) / 100).clip(0, 1)
    pct_norm = (
        (pct_from_high.reindex(trending.index).fillna(-100) + 100) / 100
    ).clip(0, 1)

    score = (
        td_norm  * 0.30 +
        rsi_norm * 0.20 +
        adx_norm * 0.20 +
        rs_norm  * 0.20 +
        pct_norm * 0.10
    ) * 100

    score = score.round(1)
    score.name = "Momentum Score"
    return score


# ─────────────────────────────────────────────────────────────────────
#  NSE MASTER + SECTOR MASTER LOADERS
# ─────────────────────────────────────────────────────────────────────
def load_nse_master(path):
    try:
        master = pd.read_csv(path, dtype=str)
        master.columns = master.columns.str.strip().str.upper()
        master["SYMBOL"] = master["SYMBOL"].str.strip().str.upper()
        master = master.rename(columns={
            "NAME OF COMPANY": "Company Name",
            "ISIN NUMBER":     "ISIN",
        })
        keep = [c for c in ["SYMBOL", "Company Name", "ISIN"]
                if c in master.columns]
        master = (master[keep]
                  .drop_duplicates("SYMBOL")
                  .set_index("SYMBOL"))
        print(f"  NSE master  : {len(master)} equity symbols")
        return master
    except Exception as e:
        print(f"  [WARN] NSE master failed: {e}")
        return pd.DataFrame()


def load_sector_master(path):
    try:
        sc = pd.read_csv(path, dtype=str)
        sc.columns = sc.columns.str.strip()
        sc["Symbol"] = sc["Symbol"].str.strip().str.upper()
        sc = sc[["Symbol", "Sector", "CapCategory"]].drop_duplicates("Symbol")
        sc = sc.set_index("Symbol")
        print(f"  Sector master : {len(sc)} symbols loaded")
        return sc
    except Exception as e:
        print(f"  [WARN] Sector master failed: {e}")
        return pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────
#  EXCEL STYLES
# ─────────────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1F3864"
CLR_HEADER_FG   = "FFFFFF"
CLR_GREEN_DARK  = "1E8449"
CLR_RED_DARK    = "C0392B"
CLR_GREEN_LIGHT = "D5F5E3"
CLR_RED_LIGHT   = "FADBD8"
CLR_ROW_ODD     = "F2F2F2"
CLR_ROW_EVEN    = "FFFFFF"

# EMA signal colours
EMA_CLR_GOLDEN  = "145A32"   # dark green  — golden cross (best buy signal)
EMA_CLR_ABOVE   = "A9DFBF"   # light green — above 200 EMA (uptrend)
EMA_CLR_AMBER   = "F39C12"   # amber       — approaching (watch)
EMA_CLR_FG_DARK = "FFFFFF"   # white text on dark bg
EMA_CLR_FG_LITE = "145A32"   # dark green text on light green bg

EMA_SIGNAL_LABELS = {
    "golden_cross": "Golden Cross ✓",
    "above_200":    "Above 200 EMA",
    "approaching":  "Approaching ▲",
    "":             "",
}

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_style():
    return {
        "font":      Font(bold=True, color=CLR_HEADER_FG, name="Arial", size=9),
        "fill":      PatternFill("solid", fgColor=CLR_HEADER_BG),
        "alignment": Alignment(horizontal="center", vertical="center",
                               wrap_text=True),
        "border":    BORDER,
    }


def _apply(cell, **kwargs):
    for k, v in kwargs.items():
        setattr(cell, k, v)


def _col_widths(df):
    widths = []
    for col in df.columns:
        if col == "Company Name":     widths.append(38)
        elif col == "Symbol":         widths.append(13)
        elif col == "ISIN":           widths.append(14)
        elif col == "Sector":         widths.append(28)
        elif col == "CapCategory":    widths.append(12)
        elif col == "Trending Days":  widths.append(10)
        elif "Change%" in col:        widths.append(11)
        elif col in ("ADX 14", "RSI 14"): widths.append(9)
        elif col == "EMA Signal":     widths.append(17)
        elif col in ("EMA 50", "EMA 200"): widths.append(11)
        elif "-" in col and len(col) == 10: widths.append(11)   # date col
        else:                         widths.append(12)
    return widths


def _write_ema_signal_cell(cell, signal_val, bg):
    """Apply EMA signal formatting to a cell."""
    label = EMA_SIGNAL_LABELS.get(signal_val, "")
    cell.value = label
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER

    if signal_val == "golden_cross":
        cell.fill = PatternFill("solid", fgColor=EMA_CLR_GOLDEN)
        cell.font = Font(bold=True, color=EMA_CLR_FG_DARK, name="Arial", size=8)
    elif signal_val == "above_200":
        cell.fill = PatternFill("solid", fgColor=EMA_CLR_ABOVE)
        cell.font = Font(bold=True, color=EMA_CLR_FG_LITE, name="Arial", size=8)
    elif signal_val == "approaching":
        cell.fill = PatternFill("solid", fgColor=EMA_CLR_AMBER)
        cell.font = Font(bold=True, color=CLR_HEADER_BG, name="Arial", size=8)
    else:
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(name="Arial", size=8, color="AAAAAA")


# ─────────────────────────────────────────────────────────────────────
#  SHEET 1 — TRUE / FALSE
# ─────────────────────────────────────────────────────────────────────
def write_sheet_truefalse(ws, df, date_col_names):
    h = _header_style()
    for ci, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        _apply(cell, **h)

    cols     = list(df.columns)
    date_set = set(date_col_names)

    for ri, row in enumerate(df.itertuples(index=False), start=2):
        bg = CLR_ROW_ODD if ri % 2 == 0 else CLR_ROW_EVEN
        for ci, (col, val) in enumerate(zip(cols, row), start=1):
            cell = ws.cell(row=ri, column=ci)

            if col == "EMA Signal":
                _write_ema_signal_cell(cell, str(val) if pd.notna(val) else "", bg)

            elif col in date_set:
                if val is True or val == True:
                    cell.value = "TRUE"
                    cell.fill  = PatternFill("solid", fgColor=CLR_GREEN_DARK)
                    cell.font  = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                elif val is False or val == False:
                    cell.value = "FALSE"
                    cell.fill  = PatternFill("solid", fgColor=CLR_RED_DARK)
                    cell.font  = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                else:
                    cell.value = ""
                    cell.fill  = PatternFill("solid", fgColor="EEEEEE")
                    cell.font  = Font(name="Arial", size=8)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER

            elif col == "Trending Days":
                cell.value     = int(val) if pd.notna(val) else ""
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.font      = Font(bold=True, name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = BORDER

            elif "Change%" in col:
                if pd.notna(val):
                    cell.value         = float(val) / 100
                    cell.number_format = "+0.00%;-0.00%;0.00%"
                    cell.fill  = PatternFill(
                        "solid",
                        fgColor=CLR_GREEN_LIGHT if val >= 0 else CLR_RED_LIGHT
                    )
                    cell.font  = Font(
                        bold=True,
                        color=CLR_GREEN_DARK if val >= 0 else CLR_RED_DARK,
                        name="Arial", size=9,
                    )
                else:
                    cell.value = ""
                    cell.fill  = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = BORDER

            elif col in ("ADX 14", "RSI 14", "EMA 50", "EMA 200"):
                cell.value         = float(val) if pd.notna(val) else ""
                cell.number_format = "0.00"
                cell.fill          = PatternFill("solid", fgColor=bg)
                cell.font          = Font(name="Arial", size=9)
                cell.alignment     = Alignment(horizontal="center", vertical="center")
                cell.border        = BORDER

            else:
                cell.value     = val if pd.notna(val) else ""
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.font      = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=False)
                cell.border    = BORDER

    for ci, w in enumerate(_col_widths(df), start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────
#  SHEET 2 — % CHANGE DAILY
# ─────────────────────────────────────────────────────────────────────
def write_sheet_pct(ws, df_base, pct_matrix, date_col_names):
    df = df_base.copy()
    for dc in date_col_names:
        if dc in pct_matrix.columns:
            df[dc] = pct_matrix[dc].values

    h        = _header_style()
    cols     = list(df.columns)
    date_set = set(date_col_names)

    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        _apply(cell, **h)

    for ri, row in enumerate(df.itertuples(index=False), start=2):
        bg = CLR_ROW_ODD if ri % 2 == 0 else CLR_ROW_EVEN
        for ci, (col, val) in enumerate(zip(cols, row), start=1):
            cell = ws.cell(row=ri, column=ci)

            if col == "EMA Signal":
                _write_ema_signal_cell(cell, str(val) if pd.notna(val) else "", bg)

            elif col in date_set:
                if pd.notna(val):
                    pct_val            = float(val)
                    cell.value         = pct_val / 100
                    cell.number_format = '+0.00%;-0.00%;"-"'
                    if pct_val > 0:
                        cell.fill = PatternFill("solid", fgColor=CLR_GREEN_DARK)
                        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                    elif pct_val < 0:
                        cell.fill = PatternFill("solid", fgColor=CLR_RED_DARK)
                        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                    else:
                        cell.fill = PatternFill("solid", fgColor="DDDDDD")
                        cell.font = Font(name="Arial", size=8)
                else:
                    cell.value = ""
                    cell.fill  = PatternFill("solid", fgColor="EEEEEE")
                    cell.font  = Font(name="Arial", size=8)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = BORDER

            elif col == "Trending Days":
                cell.value     = int(val) if pd.notna(val) else ""
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.font      = Font(bold=True, name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = BORDER

            elif "Change%" in col:
                if pd.notna(val):
                    cell.value         = float(val) / 100
                    cell.number_format = "+0.00%;-0.00%;0.00%"
                    cell.fill  = PatternFill(
                        "solid",
                        fgColor=CLR_GREEN_LIGHT if float(val) >= 0 else CLR_RED_LIGHT
                    )
                    cell.font  = Font(
                        bold=True,
                        color=CLR_GREEN_DARK if float(val) >= 0 else CLR_RED_DARK,
                        name="Arial", size=9,
                    )
                else:
                    cell.value = ""
                    cell.fill  = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = BORDER

            elif col in ("ADX 14", "RSI 14", "EMA 50", "EMA 200"):
                cell.value         = float(val) if pd.notna(val) else ""
                cell.number_format = "0.00"
                cell.fill          = PatternFill("solid", fgColor=bg)
                cell.font          = Font(name="Arial", size=9)
                cell.alignment     = Alignment(horizontal="center", vertical="center")
                cell.border        = BORDER

            else:
                cell.value     = val if pd.notna(val) else ""
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.font      = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = BORDER

    for ci, w in enumerate(_col_widths(df), start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────
#  SECTOR AGGREGATION
# ─────────────────────────────────────────────────────────────────────
def build_sector_12d(result):
    df  = result.dropna(subset=["Sector"]).copy()
    grp = df.groupby("Sector", sort=False)
    agg = pd.DataFrame({
        "Stocks":            grp["Symbol"].count(),
        "Avg Trending Days": grp["Trending Days"].mean().round(1),
        "Avg 12d Change%":   grp["12d Change%"].mean().round(2),
        "Avg RSI 14":        grp["RSI 14"].mean().round(2),
        "Avg ADX 14":        grp["ADX 14"].mean().round(2),
    }).reset_index()
    return agg.sort_values("Avg Trending Days", ascending=False).reset_index(drop=True)


def build_sector_5d(result):
    df  = result.dropna(subset=["Sector"]).copy()
    grp = df.groupby("Sector", sort=False)
    agg = pd.DataFrame({
        "Stocks":          grp["Symbol"].count(),
        "Avg 5d Change%":  grp["5d Change%"].mean().round(2),
        "Avg RSI 14":      grp["RSI 14"].mean().round(2),
        "Avg ADX 14":      grp["ADX 14"].mean().round(2),
    }).reset_index()
    return agg.sort_values("Avg 5d Change%", ascending=False).reset_index(drop=True)


def build_sector_today(result, today_col, pct_today_series=None):
    df = result.dropna(subset=["Sector"]).copy()
    df["_up"]   = df[today_col].apply(
        lambda x: 1 if x is True or x == True else 0
    )
    df["_down"] = df[today_col].apply(
        lambda x: 1 if x is False or x == False else 0
    )
    if pct_today_series is not None:
        df["_pct_today"] = df["Symbol"].map(pct_today_series)

    grp = df.groupby("Sector", sort=False)
    cols = {
        "Stocks":      grp["Symbol"].count(),
        "Stocks Up":   grp["_up"].sum().astype(int),
        "Stocks Down": grp["_down"].sum().astype(int),
    }
    if pct_today_series is not None:
        cols["Avg Today Change%"] = grp["_pct_today"].mean().round(2)

    agg = pd.DataFrame(cols).reset_index()
    agg["% Stocks Up"] = (agg["Stocks Up"] / agg["Stocks"] * 100).round(1)
    if pct_today_series is not None:
        agg = agg[["Sector", "Stocks", "Stocks Up", "Stocks Down",
                   "% Stocks Up", "Avg Today Change%"]]
    return agg.sort_values("% Stocks Up", ascending=False).reset_index(drop=True)


# keep old name for engine_db.py compatibility
def build_sector_today_with_pct(result, today_col, pct_today_series):
    return build_sector_today(result, today_col, pct_today_series)


# ─────────────────────────────────────────────────────────────────────
#  SECTOR SHEET WRITER
# ─────────────────────────────────────────────────────────────────────
CLR_RANK_TOP = "1E8449"
CLR_RANK_MID = "F39C12"
CLR_RANK_BOT = "C0392B"


def _rank_fill(rank, total):
    third = max(1, total // 3)
    if rank < third:       return CLR_RANK_TOP
    elif rank < 2 * third: return CLR_RANK_MID
    else:                  return CLR_RANK_BOT


def write_sector_sheet(ws, df, pct_cols=None, rank_col=None):
    if pct_cols is None:
        pct_cols = []
    h     = _header_style()
    cols  = list(df.columns)
    total = len(df)

    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        _apply(cell, **h)

    for ri, row in enumerate(df.itertuples(index=False), start=2):
        rank   = ri - 2
        accent = _rank_fill(rank, total)
        bg     = CLR_ROW_ODD if ri % 2 == 0 else CLR_ROW_EVEN

        for ci, (col, val) in enumerate(zip(cols, row), start=1):
            cell = ws.cell(row=ri, column=ci)

            if ci == 1:
                cell.fill      = PatternFill("solid", fgColor=accent)
                cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                cell.value     = val if pd.notna(val) else ""
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col == "Stocks":
                cell.value     = int(val) if pd.notna(val) else ""
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.font      = Font(bold=True, name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == "Stocks Up":
                cell.value     = int(val) if pd.notna(val) else ""
                cell.fill      = PatternFill("solid", fgColor="D5F5E3")
                cell.font      = Font(bold=True, color=CLR_GREEN_DARK, name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == "Stocks Down":
                cell.value     = int(val) if pd.notna(val) else ""
                cell.fill      = PatternFill("solid", fgColor="FADBD8")
                cell.font      = Font(bold=True, color=CLR_RED_DARK, name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == "% Stocks Up":
                v              = float(val) if pd.notna(val) else 0.0
                cell.value     = v / 100
                cell.number_format = "0.0%"
                cell.fill      = PatternFill(
                    "solid",
                    fgColor=CLR_GREEN_LIGHT if v >= 50 else CLR_RED_LIGHT
                )
                cell.font      = Font(
                    bold=True,
                    color=CLR_GREEN_DARK if v >= 50 else CLR_RED_DARK,
                    name="Arial", size=9,
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col in pct_cols:
                if pd.notna(val):
                    v              = float(val)
                    cell.value     = v / 100
                    cell.number_format = "+0.00%;-0.00%;0.00%"
                    cell.fill      = PatternFill(
                        "solid",
                        fgColor=CLR_GREEN_LIGHT if v >= 0 else CLR_RED_LIGHT
                    )
                    cell.font      = Font(
                        bold=True,
                        color=CLR_GREEN_DARK if v >= 0 else CLR_RED_DARK,
                        name="Arial", size=9,
                    )
                else:
                    cell.value = ""
                    cell.fill  = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value         = float(val) if pd.notna(val) else ""
                cell.number_format = "0.00"
                cell.fill          = PatternFill("solid", fgColor=bg)
                cell.font          = Font(name="Arial", size=9)
                cell.alignment     = Alignment(horizontal="center", vertical="center")

            cell.border = BORDER

    for ci, col in enumerate(cols, start=1):
        if col == "Sector":                  w = 30
        elif col in ("Stocks Up", "Stocks Down", "% Stocks Up"): w = 13
        elif "Change%" in col or "Today" in col: w = 16
        elif col in ("Avg RSI 14", "Avg ADX 14"): w = 12
        elif col == "Avg Trending Days":     w = 16
        elif col == "Stocks":                w = 9
        else:                                w = 14
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═" * 55)
    print("   NSE Momentum Analytics Engine  (EQ + BE)")
    print("═" * 55 + "\n")

    # 1. Load bhav files
    print("STEP 1 — Loading bhav files (EQ + BE)...")
    raw, isin_map = load_bhav_files(DATA_FOLDER)
    total_dates = raw["DATE"].nunique()
    if total_dates < LOOKBACK_DAYS + 1:
        raise RuntimeError(
            f"Need ≥{LOOKBACK_DAYS + 1} trading days. Found {total_dates}."
        )

    # 2. Pivots
    print("STEP 2 — Building price pivots...")
    close_piv, high_piv, low_piv = build_price_pivots(raw)
    print(f"  Pivot : {close_piv.shape[0]} symbols × {close_piv.shape[1]} dates\n")

    # 3. Matrix dates
    print("STEP 3 — 12-day rolling window...")
    matrix_dates = get_matrix_dates(close_piv, LOOKBACK_DAYS)

    # 4. True/False + daily %
    print("STEP 4 — Computing matrices...")
    trending, bool_matrix, pct_matrix = compute_matrices(close_piv, matrix_dates)

    # 5. Period changes
    print("STEP 5 — Computing 12d and 5d Change%...")
    chg_12d, chg_5d = compute_period_changes(close_piv, matrix_dates)

    # 6. RSI
    print(f"STEP 6 — RSI({RSI_PERIOD})...")
    rsi = compute_rsi(close_piv, RSI_PERIOD)
    print(f"  RSI populated : {rsi.notna().sum()} symbols\n")

    # 7. ADX
    print(f"STEP 7 — ADX({ADX_PERIOD})...")
    adx = compute_adx(close_piv, high_piv, low_piv, ADX_PERIOD, MIN_ROWS_ADX)
    print(f"  ADX populated : {adx.notna().sum()} symbols\n")

    # 8. EMA 50 / 200 + Golden Cross signal
    print(f"STEP 8 — EMA({EMA_FAST}) / EMA({EMA_SLOW}) + Golden Cross signal...")
    ema_df = compute_ema_signals(
        close_piv,
        fast=EMA_FAST,
        slow=EMA_SLOW,
        cross_window=EMA_CROSS_WINDOW,
        min_rows=EMA_MIN_ROWS,
    )
    signal_counts = ema_df["EMA Signal"].value_counts()
    print(f"  Golden Cross  : {signal_counts.get('golden_cross', 0)}")
    print(f"  Above 200 EMA : {signal_counts.get('above_200', 0)}")
    print(f"  Approaching   : {signal_counts.get('approaching', 0)}")
    print(f"  No signal     : {signal_counts.get('', 0)}\n")

    # 9. Column order: newest → oldest
    date_cols_nf   = list(reversed(matrix_dates))
    date_col_names = [d.strftime("%Y-%m-%d") for d in date_cols_nf]
    bool_out = bool_matrix[date_cols_nf].copy()
    bool_out.columns = date_col_names
    pct_out  = pct_matrix[date_cols_nf].copy()
    pct_out.columns  = date_col_names

    # 10. Assemble base dataframe
    print("STEP 9 — Assembling base dataframe...")
    result = (
        pd.DataFrame({"Trending Days": trending})
        .join(chg_12d, how="left")
        .join(chg_5d,  how="left")
        .join(bool_out, how="left")
        .join(adx,      how="left")
        .join(rsi,      how="left")
        .join(ema_df,   how="left")   # ← EMA 50, EMA 200, EMA Signal
    )
    result.index.name = "Symbol"
    result = result.reset_index()
    result["ISIN"] = result["Symbol"].map(isin_map)
    print(f"  Base dataframe : {len(result)} symbols\n")

    # 11. NSE master
    print("STEP 10 — Merging NSE master...")
    if NSE_MASTER_CSV:
        master = load_nse_master(NSE_MASTER_CSV)
        if not master.empty:
            before = len(result)
            result = result.drop(columns=["ISIN"], errors="ignore")
            result = result.join(master, on="Symbol", how="inner")
            print(f"  {before} → {len(result)} symbols (ETFs excluded)\n")

    # 12. Sector master
    print("STEP 11 — Merging Sector + Cap master...")
    if NSE_SECTOR_MASTER:
        sc = load_sector_master(NSE_SECTOR_MASTER)
        if not sc.empty:
            result = result.join(sc, on="Symbol", how="left")
            filled = result["Sector"].notna().sum()
            print(f"  Sector filled : {filled}/{len(result)} symbols\n")

    # 13. Sort
    result = result.sort_values(
        ["Trending Days", "RSI 14", "ADX 14"],
        ascending=[False, False, False]
    ).reset_index(drop=True)

    # 14. Final column order
    ordered = (
        ["Company Name", "Symbol", "ISIN", "Sector", "CapCategory",
         "Trending Days", "12d Change%", "5d Change%"] +
        date_col_names +
        ["ADX 14", "RSI 14", "EMA 50", "EMA 200", "EMA Signal"]
    )
    final_cols = [c for c in ordered if c in result.columns]
    result = result[final_cols]

    # 15. Write Excel
    print("STEP 12 — Writing Excel workbook...")
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Trend (True-False)"
    write_sheet_truefalse(ws1, result, date_col_names)

    ws2 = wb.create_sheet("Trend (% Change)")
    pct_aligned = (pct_out
                   .loc[pct_out.index.isin(result["Symbol"])]
                   .reindex(result["Symbol"].values))
    write_sheet_pct(ws2, result, pct_aligned, date_col_names)

    today_col = date_col_names[0]

    ws3 = wb.create_sheet("Sector Momentum (12d)")
    sec_12d = build_sector_12d(result)
    write_sector_sheet(ws3, sec_12d, pct_cols=["Avg 12d Change%"])
    print(f"  Sheet 3 : {len(sec_12d)} sectors (12d)")

    ws4 = wb.create_sheet("Sector Momentum (5d)")
    sec_5d = build_sector_5d(result)
    write_sector_sheet(ws4, sec_5d, pct_cols=["Avg 5d Change%"])
    print(f"  Sheet 4 : {len(sec_5d)} sectors (5d)")

    ws5 = wb.create_sheet("Sector Momentum (Today)")
    pct_today_map = dict(zip(result["Symbol"].values,
                             pct_aligned[today_col].values
                             if today_col in pct_aligned.columns
                             else [np.nan] * len(result)))
    sec_today = build_sector_today(result, today_col,
                                   pd.Series(pct_today_map))
    write_sector_sheet(ws5, sec_today, pct_cols=["Avg Today Change%"])
    print(f"  Sheet 5 : {len(sec_today)} sectors (today)\n")

    wb.save(OUTPUT_FILE)

    print(f"\n{'═' * 55}")
    print(f"  ✅  Saved    : {OUTPUT_FILE}")
    print(f"  Symbols     : {len(result)}")
    print(f"  TODAY col   : {date_col_names[0]}")
    print(f"{'═' * 55}")

    # Preview
    preview_cols = ["Company Name", "Symbol", "Sector",
                    "Trending Days", "12d Change%",
                    "ADX 14", "RSI 14", "EMA Signal"]
    preview_cols = [c for c in preview_cols if c in result.columns]
    print("\nTop 10 by Trending Days:")
    print(result[preview_cols].head(10).to_string(index=False))

    golden = result[result["EMA Signal"] == "golden_cross"]
    if not golden.empty:
        print(f"\nGolden Cross stocks ({len(golden)}):")
        print(golden[preview_cols].head(10).to_string(index=False))


if __name__ == "__main__":
    main()