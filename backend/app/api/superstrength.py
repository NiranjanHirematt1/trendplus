"""
GET /api/superstrength          — Super Strength screen
GET /api/superstrength/excel    — Download today's 12-day matrix Excel
"""
import datetime
import io
import logging
from typing import Optional

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from app.core.database import get_pool

logger = logging.getLogger(__name__)
router = APIRouter()


def _resolve_date(raw) -> Optional[datetime.date]:
    if raw is None:
        return None
    if isinstance(raw, datetime.date):
        return raw
    try:
        return datetime.date.fromisoformat(str(raw))
    except ValueError:
        return None


@router.get("", summary="Super Strength screen")
async def get_superstrength(
    tab:     str           = Query("details"),
    date:    Optional[str] = Query(None),
    sector:  Optional[str] = Query(None),
    cap:     Optional[str] = Query(None),
    min_rpi: float         = Query(0, ge=0),
    pool=Depends(get_pool),
):
    async with pool.acquire() as conn:
        if date:
            trade_date = _resolve_date(date)
            if not trade_date:
                raise HTTPException(422, "Invalid date format")
        else:
            row = await conn.fetchrow("SELECT trade_date FROM v_latest_date")
            trade_date = _resolve_date(row["trade_date"] if row else None)
        if not trade_date:
            return {"date": None, "total": 0, "data": []}

        params     = [trade_date]
        conditions = ["tr.trade_date = $1", "s.is_active = true"]
        p = 2

        if min_rpi > 0:
            conditions.append(f"tr.weighted_rpi >= ${p}")
            params.append(min_rpi); p += 1
        if sector:
            conditions.append(f"s.sector = ${p}")
            params.append(sector); p += 1
        if cap:
            conditions.append(f"s.cap_category = ${p}")
            params.append(cap); p += 1

        where = " AND ".join(conditions)

        # Build SELECT with safe column access — columns may not exist yet
        rows = await conn.fetch(
            f"""
            SELECT
                s.symbol, s.company_name, s.sector, s.cap_category, s.isin,
                tr.close_price      AS latest_price,
                tr.high_52w, tr.pct_from_high, tr.near_52w_high, tr.rank_52w,
                tr.rsi_14, tr.adx_14, tr.chg_1d, tr.chg_12d,
                tr.ema_50, tr.ema_200, tr.ema_signal,
                tr.momentum_score, tr.rs_score,
                -- New columns (NULL if migration not run yet)
                tr.weighted_rpi,
                tr.rpi_2w, tr.rpi_3m, tr.rpi_6m, tr.rpi_6m_sma2w,
                tr.rsi_1d, tr.rsi_1w,
                tr.ema_9, tr.ema_21
            FROM trend_results tr
            JOIN symbols s ON s.symbol = tr.symbol
            WHERE {where}
            ORDER BY s.sector ASC NULLS LAST,
                     tr.weighted_rpi DESC NULLS LAST
            """,
            *params,
        )

    data = [dict(r) for r in rows]
    for d in data:
        rank = d.get("rank_52w") or 0
        wrpi = d.get("weighted_rpi") or 0
        d["rocket_52w_wrpi"] = rank >= 95 and wrpi >= 90
        d["rocket_52w"]      = rank >= 98

    return {"date": str(trade_date), "tab": tab, "total": len(data), "data": data}


@router.get("/excel", summary="Download today's 12-day matrix Excel")
async def download_excel(
    date: Optional[str] = Query(None),
    pool=Depends(get_pool),
):
    """
    Builds Excel directly from data already stored in trend_results.
    No re-computation needed — uses pct_matrix and bool_matrix from DB.
    Fast: typically 2-4 seconds.
    """
    async with pool.acquire() as conn:
        if date:
            trade_date = _resolve_date(date)
            if not trade_date:
                raise HTTPException(422, "Invalid date")
        else:
            row = await conn.fetchrow("SELECT trade_date FROM v_latest_date")
            trade_date = _resolve_date(row["trade_date"] if row else None)
        if not trade_date:
            raise HTTPException(404, "No trading data available")

        rows = await conn.fetch(
            """
            SELECT
                s.symbol, s.company_name, s.isin, s.sector, s.cap_category,
                tr.trending_days, tr.chg_12d, tr.chg_5d, tr.chg_1d,
                tr.rsi_14, tr.adx_14, tr.momentum_score,
                tr.ema_200, tr.close_price,
                tr.bool_matrix, tr.pct_matrix
            FROM trend_results tr
            JOIN symbols s ON s.symbol = tr.symbol
            WHERE tr.trade_date = $1 AND s.is_active = true
            ORDER BY tr.momentum_score DESC NULLS LAST
            """,
            trade_date,
        )

    if not rows:
        raise HTTPException(404, f"No data for {trade_date}")

    excel_bytes = _build_excel(rows, trade_date)
    filename = f"{trade_date}_trend_matrix.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_excel(rows, trade_date: datetime.date) -> bytes:
    """Build Excel entirely from DB rows — no re-computation."""
    import json
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_BG  = "1F3864"
    GRN_D   = "1E8449"
    GRN_DK  = "145A32"   # dark green = price above EMA 200
    RED_D   = "C0392B"
    GRN_L   = "D5F5E3"
    RED_L   = "FADBD8"
    ODD     = "F2F2F2"
    EVEN    = "FFFFFF"
    THIN    = Side(style="thin", color="CCCCCC")
    BDR     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_F   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    HDR_A   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    HDR_FL  = PatternFill("solid", fgColor=HDR_BG)

    # Collect date keys sorted oldest→newest (left to right)
    date_keys = []
    for r in rows:
        pm = r["pct_matrix"]
        if pm:
            data = pm if isinstance(pm, dict) else json.loads(pm)
            keys = sorted(data.keys(), reverse=True)[:12]
            if len(keys) > len(date_keys):
                date_keys = keys
            break

    wb = Workbook()

    def _hdr(ws, cols):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font, c.fill, c.alignment, c.border = HDR_F, HDR_FL, HDR_A, BDR
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    def _w(col):
        if col == "Company Name": return 38
        if col == "Symbol":       return 12
        if col == "ISIN":         return 14
        if col == "Sector":       return 26
        if col == "Cap":          return 10
        if "Change" in col or "%" in col: return 10
        if col in ("ADX","RSI","Trend"): return 8
        if len(col) == 5 and "-" in col: return 10  # date
        return 11

    # ── Sheet 1: Trend True/False ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Trend (True-False)"
    cols1 = (["Company Name","Symbol","ISIN","Sector","Cap",
               "Trend","12d%","5d%","1d%"] + date_keys + ["ADX","RSI"])
    _hdr(ws1, cols1)

    for ri, r in enumerate(rows, 2):
        bm = r["bool_matrix"] or {}
        if isinstance(bm, str):
            import json; bm = json.loads(bm)
        bg = ODD if ri % 2 == 0 else EVEN
        vals = [
            r["company_name"] or r["symbol"], r["symbol"],
            r["isin"] or "", r["sector"] or "", r["cap_category"] or "",
            r["trending_days"],
            r["chg_12d"], r["chg_5d"], r["chg_1d"],
        ] + [bm.get(k) for k in date_keys] + [r["adx_14"], r["rsi_14"]]

        for ci, (col, val) in enumerate(zip(cols1, vals), 1):
            cell = ws1.cell(row=ri, column=ci)
            if col in date_keys:
                if val is True:
                    cell.value = "TRUE"
                    cell.fill = PatternFill("solid", fgColor=GRN_D)
                    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                elif val is False:
                    cell.value = "FALSE"
                    cell.fill = PatternFill("solid", fgColor=RED_D)
                    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                else:
                    cell.value = ""; cell.fill = PatternFill("solid", fgColor="EEEEEE")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif "%" in col or "Change" in str(col):
                if val is not None:
                    try:
                        v = float(val)
                        cell.value = v / 100
                        cell.number_format = "+0.00%;-0.00%;0.00%"
                        cell.fill = PatternFill("solid", fgColor=GRN_L if v >= 0 else RED_L)
                        cell.font = Font(bold=True, color=GRN_D if v >= 0 else RED_D, name="Arial", size=9)
                    except: cell.value = val
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = int(val) if col == "Trend" and val is not None else (val or "")
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(
                    horizontal="center" if col in ("Trend","ADX","RSI") else "left",
                    vertical="center"
                )
            cell.border = BDR

    for ci, col in enumerate(cols1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = _w(col)

    # ── Sheet 2: Trend % Change ────────────────────────────────────
    ws2 = wb.create_sheet("Trend (% Change)")
    _hdr(ws2, cols1)

    for ri, r in enumerate(rows, 2):
        pm = r["pct_matrix"] or {}
        if isinstance(pm, str):
            import json; pm = json.loads(pm)
        above_ema = (r["ema_200"] and r["close_price"] and
                     float(r["close_price"]) > float(r["ema_200"]))
        bg = ODD if ri % 2 == 0 else EVEN
        vals = [
            r["company_name"] or r["symbol"], r["symbol"],
            r["isin"] or "", r["sector"] or "", r["cap_category"] or "",
            r["trending_days"],
            r["chg_12d"], r["chg_5d"], r["chg_1d"],
        ] + [pm.get(k) for k in date_keys] + [r["adx_14"], r["rsi_14"]]

        for ci, (col, val) in enumerate(zip(cols1, vals), 1):
            cell = ws2.cell(row=ri, column=ci)
            if col in date_keys:
                if val is not None:
                    try:
                        n = float(val)
                        cell.value = n / 100
                        cell.number_format = "+0.00%;-0.00%;\"-\""
                        if n > 0:
                            fg = GRN_DK if above_ema else GRN_D
                        elif n < 0:
                            fg = RED_D
                        else:
                            fg = "DDDDDD"
                        cell.fill = PatternFill("solid", fgColor=fg)
                        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                    except:
                        cell.value = ""
                        cell.fill = PatternFill("solid", fgColor="EEEEEE")
                else:
                    cell.value = ""
                    cell.fill = PatternFill("solid", fgColor="EEEEEE")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif "%" in col or "Change" in str(col):
                if val is not None:
                    try:
                        v = float(val)
                        cell.value = v / 100
                        cell.number_format = "+0.00%;-0.00%;0.00%"
                        cell.fill = PatternFill("solid", fgColor=GRN_L if v >= 0 else RED_L)
                        cell.font = Font(bold=True, color=GRN_D if v >= 0 else RED_D, name="Arial", size=9)
                    except: cell.value = val
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = int(val) if col == "Trend" and val is not None else (val or "")
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(
                    horizontal="center" if col in ("Trend","ADX","RSI") else "left",
                    vertical="center"
                )
            cell.border = BDR

    for ci, col in enumerate(cols1, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = _w(col)

    # ── Sheet 3: Sector Momentum ───────────────────────────────────
    ws3 = wb.create_sheet("Sector Momentum")
    sec_data: dict = {}
    for r in rows:
        sec = r["sector"] or "(No Sector)"
        if sec not in sec_data:
            sec_data[sec] = {"count": 0, "trend_sum": 0, "chg12_sum": 0,
                             "rsi_sum": 0, "adx_sum": 0, "n_rsi": 0, "n_adx": 0}
        d = sec_data[sec]
        d["count"] += 1
        if r["trending_days"] is not None: d["trend_sum"] += r["trending_days"]
        if r["chg_12d"] is not None:       d["chg12_sum"] += float(r["chg_12d"])
        if r["rsi_14"] is not None:        d["rsi_sum"]  += float(r["rsi_14"]);  d["n_rsi"] += 1
        if r["adx_14"] is not None:        d["adx_sum"]  += float(r["adx_14"]);  d["n_adx"] += 1

    sec_cols = ["Sector","Stocks","Avg Trend","Avg 12d%","Avg RSI","Avg ADX"]
    _hdr(ws3, sec_cols)
    for ri, (sec, d) in enumerate(
        sorted(sec_data.items(), key=lambda x: -x[1]["trend_sum"]/max(x[1]["count"],1)), 2
    ):
        n = d["count"]
        row_vals = [
            sec, n,
            round(d["trend_sum"]/n, 1),
            round(d["chg12_sum"]/n, 2),
            round(d["rsi_sum"]/d["n_rsi"], 2) if d["n_rsi"] else "",
            round(d["adx_sum"]/d["n_adx"], 2) if d["n_adx"] else "",
        ]
        bg = ODD if ri % 2 == 0 else EVEN
        for ci, val in enumerate(row_vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=val or "")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                       vertical="center")
            cell.border = BDR
    ws3.column_dimensions["A"].width = 28
    for ci in range(2, 7):
        ws3.column_dimensions[get_column_letter(ci)].width = 12
    ws3.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()