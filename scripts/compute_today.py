#!/usr/bin/env python3
"""
scripts/compute_today.py
─────────────────────────
Computes ALL indicators for a single trading date and upserts to Supabase.

KEY DIFFERENCES from compute_all_dates.py:
  ✓ Only computes the LATEST date (or --date override)
  ✓ Does NOT store bool_matrix or pct_matrix (removed from schema)
  ✓ bool_matrix/pct_matrix only computed for the Excel download
  ✓ Handles 525+ bhav files: restricts to last 252 days (1 year)
  ✓ All Super Strength columns computed: EMA9, EMA21, RPI 2W/3M/6M, RSI 1D/1W
  ✓ Recomputation-safe: TRUNCATE + run = clean fresh data

Usage
─────
  python scripts/compute_today.py                # process today
  python scripts/compute_today.py --date 2026-03-20
  python scripts/compute_today.py --recompute    # wipe all + recompute today only

Full reset workflow:
  1. Run migration_v3.sql in Supabase (adds columns)
  2. python scripts/compute_today.py --recompute
"""
import argparse
import asyncio
import datetime
import io
import json
import logging
import os
import sys
import time
from pathlib import Path
from typing import Optional

ROOT    = Path(__file__).resolve().parent.parent
BACKEND = ROOT / "backend"
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(BACKEND))

from dotenv import load_dotenv
load_dotenv(BACKEND / ".env")

import asyncpg
import numpy as np
import pandas as pd
from backend.app.core.sector_mapping import normalize_sector_name

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("compute_today")

from analytics_engine import (
    build_price_pivots,
    get_matrix_dates,
    compute_matrices,
    compute_period_changes,
    compute_52w_high,
    compute_rs_score,
    compute_weighted_rpi,
    compute_rpi_periods,
    compute_rsi,
    compute_rsi_short,
    compute_adx,
    compute_ema_signals,
    compute_ema_short,
    compute_momentum_score,
    build_sector_12d,
    build_sector_5d,
    build_sector_today,
    LOOKBACK_DAYS,
    RSI_PERIOD,
    ADX_PERIOD,
    MIN_ROWS_ADX,
    W52_DAYS,
    EMA_FAST,
    EMA_SLOW,
    EMA_CROSS_WINDOW,
    EMA_MIN_ROWS,
)

# ── MACD ────────────────────────────────────────────────────────────
MACD_FAST     = 12
MACD_SLOW     = 26
MACD_SIGNAL_P = 9

# ── Dataset size limit ───────────────────────────────────────────────
# If price_history has > MAX_DATES trading days, restrict to last YEAR_DAYS.
# This prevents indicator drift on very long histories where early data
# has different volatility characteristics.
MAX_DATES  = 525   # above this → restrict
YEAR_DAYS  = 262   # ~1 trading year + buffer for 52W computations

_BATCH = 500


# ════════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════════
def _sf(v) -> Optional[float]:
    try:
        x = float(v)
        return None if x != x else x
    except (TypeError, ValueError):
        return None

def _si(v) -> Optional[int]:
    try:
        return int(v)
    except (TypeError, ValueError):
        return None

def _sb(v) -> Optional[bool]:
    if v is None or (isinstance(v, float) and v != v):
        return None
    return bool(v)

def _clamp(v, lo, hi) -> Optional[float]:
    x = _sf(v)
    return None if x is None else max(lo, min(hi, x))


def compute_macd(close_piv: pd.DataFrame) -> pd.DataFrame:
    prices    = close_piv.T
    ema_fast  = prices.ewm(span=MACD_FAST,     adjust=False).mean()
    ema_slow  = prices.ewm(span=MACD_SLOW,     adjust=False).mean()
    macd_line = ema_fast - ema_slow
    macd_sig  = macd_line.ewm(span=MACD_SIGNAL_P, adjust=False).mean()
    macd_hist = macd_line - macd_sig
    last = close_piv.shape[1] - 1
    out  = pd.DataFrame({
        "MACD Line":   macd_line.iloc[last],
        "MACD Signal": macd_sig.iloc[last],
        "MACD Hist":   macd_hist.iloc[last],
    }, index=close_piv.index)
    min_rows = MACD_SLOW + MACD_SIGNAL_P
    for sym in out.index:
        if int(close_piv.loc[sym].notna().sum()) < min_rows:
            out.at[sym, "MACD Line"]   = None
            out.at[sym, "MACD Signal"] = None
            out.at[sym, "MACD Hist"]   = None
    return out

# ════════════════════════════════════════════════════════════════════
#  BHAV CLEANUP — keep only latest 255 trading dates
# ════════════════════════════════════════════════════════════════════
MAX_BHAV_DATES = 255  # 1 trading year ≈ 252, buffer of 3

async def cleanup_old_bhav_dates(conn):
    """
    Check how many distinct trading dates exist in price_history.
    If > MAX_BHAV_DATES, delete the oldest rows beyond that limit.
    """
    total = await conn.fetchval(
        "SELECT COUNT(DISTINCT trade_date) FROM price_history"
    )
    logger.info("Bhav files in price_history: %d", total)

    if total <= MAX_BHAV_DATES:
        logger.info("✓ Bhav count %d ≤ %d — no cleanup needed", total, MAX_BHAV_DATES)
        return

    excess = total - MAX_BHAV_DATES
    logger.info("⚠ Excess bhav dates: %d — deleting oldest %d...", total, excess)

    # Find the cutoff date: keep only dates ranked in top MAX_BHAV_DATES (newest)
    cutoff_date = await conn.fetchval(
        """
        SELECT trade_date FROM (
            SELECT DISTINCT trade_date
            FROM price_history
            ORDER BY trade_date DESC
            LIMIT $1
        ) sub
        ORDER BY trade_date ASC
        LIMIT 1
        """,
        MAX_BHAV_DATES,
    )

    deleted = await conn.execute(
        "DELETE FROM price_history WHERE trade_date < $1",
        cutoff_date,
    )
    logger.info("✓ Deleted bhav rows older than %s  [%s]", cutoff_date, deleted)
# ════════════════════════════════════════════════════════════════════
#  DB HELPERS
# ════════════════════════════════════════════════════════════════════
async def load_price_history(conn, trade_date: datetime.date) -> pd.DataFrame:
    """
    Load price history from DB for indicator calculation.
    Applies per-day liquidity filter (total_trades >= 5000).
    Restricts to last YEAR_DAYS if more than MAX_DATES available.
    """
    # Count total distinct trading dates
    total_dates = await conn.fetchval(
        "SELECT COUNT(DISTINCT trade_date) FROM price_history"
    )
    logger.info("Total trading dates in price_history: %d", total_dates)

    restrict = total_dates > MAX_DATES
    if restrict:
        logger.info(
            "Dataset > %d dates → restricting to last %d days for accuracy",
            MAX_DATES, YEAR_DAYS
        )
        limit_clause = f"""
            WITH date_window AS (
                SELECT trade_date FROM (
                    SELECT DISTINCT trade_date FROM price_history
                    ORDER BY trade_date DESC LIMIT {YEAR_DAYS}
                ) sub ORDER BY trade_date
            )
        """
        where_extra = "AND ph.trade_date IN (SELECT trade_date FROM date_window)"
    else:
        limit_clause = ""
        where_extra  = ""

    sql = f"""
        {limit_clause}
        SELECT ph.symbol      AS "SYMBOL",
               ph.trade_date  AS "DATE",
               ph.high_price  AS "HIGH",
               ph.low_price   AS "LOW",
               ph.close_price AS "CLOSE",
               ph.volume      AS "TOTTRDQTY",
               ph.total_trades AS "TOTALTRADES"
        FROM price_history ph
        JOIN symbols s ON s.symbol = ph.symbol
        WHERE ph.close_price IS NOT NULL
          AND ph.total_trades >= 5000
          {where_extra}
        ORDER BY ph.trade_date ASC
    """
    rows = await conn.fetch(sql)
    if not rows:
        raise RuntimeError("price_history is empty — run backfill first")

    df = pd.DataFrame([dict(r) for r in rows])
    df["DATE"] = pd.to_datetime(df["DATE"])
    for col in ["HIGH", "LOW", "CLOSE"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in ["TOTTRDQTY", "TOTALTRADES"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    logger.info(
        "Loaded: %d rows, %d symbols, %d dates (restricted=%s)",
        len(df), df["SYMBOL"].nunique(), df["DATE"].nunique(), restrict
    )
    return df


async def load_master_from_db(conn):
    """Load EQUITY_L and SECTOR_MASTER from master_data table."""
    rows = await conn.fetch("SELECT key, content FROM master_data")
    master = {r["key"]: r["content"] for r in rows}

    nse_master = pd.DataFrame()
    sec_master = pd.DataFrame()

    if "EQUITY_L" in master:
        try:
            df = pd.read_csv(io.StringIO(master["EQUITY_L"]), dtype=str)
            df.columns = [c.strip().upper() for c in df.columns]
            sym_col  = next((c for c in df.columns if "SYMBOL" in c), None)
            name_col = next((c for c in df.columns if "NAME" in c), None)
            isin_col = next((c for c in df.columns if "ISIN" in c), None)
            if sym_col:
                df = df.rename(columns={sym_col: "Symbol"})
                if name_col: df = df.rename(columns={name_col: "Company Name"})
                if isin_col: df = df.rename(columns={isin_col: "ISIN"})
                cols = ["Symbol"] + [c for c in ["Company Name","ISIN"] if c in df.columns]
                nse_master = df[cols].dropna(subset=["Symbol"]).copy()
                nse_master["Symbol"] = nse_master["Symbol"].str.strip().str.upper()
                nse_master = nse_master.set_index("Symbol")
                logger.info("EQUITY_L: %d symbols", len(nse_master))
        except Exception as e:
            logger.warning("EQUITY_L parse error: %s", e)

    if "SECTOR_MASTER" in master:
        try:
            df = pd.read_csv(io.StringIO(master["SECTOR_MASTER"]), dtype=str)
            df.columns = [c.strip().upper() for c in df.columns]
            sym_col = next((c for c in df.columns if "SYMBOL" in c), None)
            sec_col = next((c for c in df.columns if "SECTOR" in c), None)
            cap_col = next((c for c in df.columns if "CAP" in c), None)
            if sym_col and sec_col:
                rename = {sym_col: "Symbol", sec_col: "Sector"}
                if cap_col: rename[cap_col] = "CapCategory"
                df = df.rename(columns=rename)
                keep = ["Symbol", "Sector"] + (["CapCategory"] if cap_col else [])
                sec_master = df[keep].dropna(subset=["Symbol"]).copy()
                sec_master["Symbol"] = sec_master["Symbol"].str.strip().str.upper()
                sec_master["Sector"] = sec_master["Sector"].map(normalize_sector_name)
                sec_master = sec_master.set_index("Symbol")
                logger.info("SECTOR_MASTER: %d symbols", len(sec_master))
        except Exception as e:
            logger.warning("SECTOR_MASTER parse error: %s", e)

    return nse_master, sec_master


# ════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT — 12-day matrix for today
# ════════════════════════════════════════════════════════════════════
def build_today_excel(
    result: pd.DataFrame,
    bool_matrix: pd.DataFrame,
    pct_matrix:  pd.DataFrame,
    date_col_names: list[str],
    trade_date: datetime.date,
) -> bytes:
    """
    Build Excel workbook in memory with 12-day matrix for today.
    Returns bytes ready to send as HTTP response.

    Sheets:
      1. Trend (True-False)  — bool matrix per day
      2. Trend (% Change)    — pct change per day (green/red)
      3. Sector Momentum     — aggregated sector data
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    CLR_HDR_BG    = "1F3864"
    CLR_GREEN_D   = "1E8449"
    CLR_RED_D     = "C0392B"
    CLR_GREEN_L   = "D5F5E3"
    CLR_RED_L     = "FADBD8"
    CLR_ROW_ODD   = "F2F2F2"
    CLR_ROW_EVEN  = "FFFFFF"
    THIN  = Side(style="thin", color="CCCCCC")
    BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_F = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    HDR_A = Alignment(horizontal="center", vertical="center", wrap_text=True)
    HDR_FILL = PatternFill("solid", fgColor=CLR_HDR_BG)

    # Build base dataframe
    cols_keep = (
        ["Company Name", "Symbol", "ISIN", "Sector", "CapCategory",
         "Trending Days", "12d Change%", "5d Change%"]
        + date_col_names
        + ["ADX 14", "RSI 14"]
    )
    base = result[[c for c in cols_keep if c in result.columns]].copy()

    def _write_header(ws, df):
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font, c.fill, c.alignment, c.border = HDR_F, HDR_FILL, HDR_A, BDR
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"

    def _col_w(col):
        if col == "Company Name": return 38
        if col == "Symbol":       return 13
        if col == "ISIN":         return 14
        if col == "Sector":       return 28
        if col == "CapCategory":  return 12
        if col == "Trending Days": return 10
        if "Change%" in col:      return 11
        if col in ("ADX 14","RSI 14"): return 9
        if len(col) == 10 and "-" in col: return 11
        return 12

    date_set = set(date_col_names)

    wb = Workbook()

    # ── Sheet 1: True-False ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Trend (True-False)"
    _write_header(ws1, base)

    for ri, row in enumerate(base.itertuples(index=False), 2):
        bg = CLR_ROW_ODD if ri % 2 == 0 else CLR_ROW_EVEN
        for ci, (col, val) in enumerate(zip(base.columns, row), 1):
            cell = ws1.cell(row=ri, column=ci)
            if col in date_set:
                if val is True:
                    cell.value = "TRUE"
                    cell.fill = PatternFill("solid", fgColor=CLR_GREEN_D)
                    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                elif val is False:
                    cell.value = "FALSE"
                    cell.fill = PatternFill("solid", fgColor=CLR_RED_D)
                    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                else:
                    cell.value = ""
                    cell.fill = PatternFill("solid", fgColor="EEEEEE")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif "Change%" in col:
                if pd.notna(val):
                    cell.value = float(val) / 100
                    cell.number_format = "+0.00%;-0.00%;0.00%"
                    cell.fill = PatternFill("solid",
                        fgColor=CLR_GREEN_L if float(val) >= 0 else CLR_RED_L)
                    cell.font = Font(bold=True,
                        color=CLR_GREEN_D if float(val) >= 0 else CLR_RED_D,
                        name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col in ("ADX 14","RSI 14","Trending Days"):
                cell.value = (int(val) if col == "Trending Days" else
                              round(float(val),2)) if pd.notna(val) else ""
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = val if pd.notna(val) else ""
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BDR

    for ci, col in enumerate(base.columns, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = _col_w(col)

    # ── Sheet 2: % Change ────────────────────────────────────────────
    ws2 = wb.create_sheet("Trend (% Change)")
    pct_base = base.copy()
    for dc in date_col_names:
        if dc in pct_matrix.columns and dc in pct_base.columns:
            pct_base[dc] = pct_matrix[dc].reindex(result["Symbol"].values).values

    _write_header(ws2, pct_base)
    for ri, row in enumerate(pct_base.itertuples(index=False), 2):
        bg = CLR_ROW_ODD if ri % 2 == 0 else CLR_ROW_EVEN
        # check if this symbol's close > ema_200
        sym = row[list(pct_base.columns).index("Symbol")] if "Symbol" in pct_base.columns else None
        ema200_val = None
        if sym is not None and "EMA 200" in result.columns:
            row_data = result[result["Symbol"] == sym]
            if not row_data.empty:
                ema200_val = _sf(row_data.iloc[0].get("EMA 200"))
        close_val = None
        if sym is not None and "Close" in result.columns:
            row_data = result[result["Symbol"] == sym]
            if not row_data.empty:
                close_val = _sf(row_data.iloc[0].get("Close"))
        above_ema200 = (ema200_val is not None and close_val is not None
                        and close_val > ema200_val)

        for ci, (col, val) in enumerate(zip(pct_base.columns, row), 1):
            cell = ws2.cell(row=ri, column=ci)
            if col in date_set:
                if pd.notna(val):
                    pv = float(val)
                    cell.value = pv / 100
                    cell.number_format = "+0.00%;-0.00%;\"-\""
                    if pv > 0:
                        # DARK green if price > EMA 200, regular green otherwise
                        fg = "145A32" if above_ema200 else CLR_GREEN_D
                        cell.fill = PatternFill("solid", fgColor=fg)
                        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                    elif pv < 0:
                        cell.fill = PatternFill("solid", fgColor=CLR_RED_D)
                        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
                    else:
                        cell.fill = PatternFill("solid", fgColor="DDDDDD")
                else:
                    cell.value = ""
                    cell.fill = PatternFill("solid", fgColor="EEEEEE")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif "Change%" in col:
                if pd.notna(val):
                    cell.value = float(val) / 100
                    cell.number_format = "+0.00%;-0.00%;0.00%"
                    cell.fill = PatternFill("solid",
                        fgColor=CLR_GREEN_L if float(val) >= 0 else CLR_RED_L)
                    cell.font = Font(bold=True,
                        color=CLR_GREEN_D if float(val) >= 0 else CLR_RED_D,
                        name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = val if pd.notna(val) else ""
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(
                    horizontal="center" if col in ("ADX 14","RSI 14","Trending Days")
                    else "left", vertical="center")
            cell.border = BDR

    for ci, col in enumerate(pct_base.columns, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = _col_w(col)

    # ── Sheet 3: Sector Momentum ─────────────────────────────────────
    ws3 = wb.create_sheet("Sector Momentum")
    today_col = date_col_names[0]
    if "Sector" in result.columns and today_col in result.columns:
        sec_12d = build_sector_12d(result)
        _write_header(ws3, sec_12d)
        for ri, row in enumerate(sec_12d.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                cell = ws3.cell(row=ri, column=ci, value=val if pd.notna(val) else "")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BDR

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════════════
#  CORE COMPUTE + UPSERT
# ════════════════════════════════════════════════════════════════════
async def compute_and_upsert_today(
    pool,
    hist:       pd.DataFrame,
    nse_master: pd.DataFrame,
    sec_master: pd.DataFrame,
    trade_date: datetime.date,
    return_excel: bool = False,
) -> dict:
    """
    Compute all indicators for trade_date and upsert to Supabase.
    Returns summary dict. If return_excel=True, also returns Excel bytes.
    """
    t0 = time.monotonic()

    # Restrict hist to this date and all history before it
    target_ts = pd.Timestamp(trade_date)
    hist_to_date = hist[hist["DATE"] <= target_ts].copy()

    if hist_to_date.empty or target_ts not in hist_to_date["DATE"].values:
        logger.warning("[%s] No price data — skipping", trade_date)
        return {"symbols": 0, "elapsed": 0}

    # Build pivots
    close_piv, high_piv, low_piv = build_price_pivots(hist_to_date)

    if close_piv.columns[-1] != target_ts:
        logger.warning("[%s] Last pivot col is %s — skipping",
                       trade_date, close_piv.columns[-1])
        return {"symbols": 0, "elapsed": 0}

    # ── Compute all indicators ────────────────────────────────────────
    matrix_dates = get_matrix_dates(close_piv, LOOKBACK_DAYS)
    trending, bool_matrix, pct_matrix = compute_matrices(close_piv, matrix_dates)
    chg_12d, chg_5d = compute_period_changes(close_piv, matrix_dates)

    high_52w, pct_from_high, near_high, rank_52w = compute_52w_high(close_piv, W52_DAYS)
    rs_score     = compute_rs_score(close_piv)
    weighted_rpi = compute_weighted_rpi(close_piv)
    rpi_periods  = compute_rpi_periods(close_piv)
    rsi          = compute_rsi(close_piv, RSI_PERIOD)
    rsi_short    = compute_rsi_short(close_piv)
    adx          = compute_adx(close_piv, high_piv, low_piv, ADX_PERIOD, MIN_ROWS_ADX)
    momentum     = compute_momentum_score(trending, rsi, adx, rs_score, pct_from_high)
    macd         = compute_macd(close_piv)
    ema_df       = compute_ema_signals(
        close_piv, fast=EMA_FAST, slow=EMA_SLOW,
        cross_window=EMA_CROSS_WINDOW, min_rows=EMA_MIN_ROWS,
    )
    ema_short    = compute_ema_short(close_piv)

    # ── Matrix for Excel only (NOT stored in DB) ──────────────────────
    date_cols_nf   = list(reversed(matrix_dates))
    date_col_names = [d.strftime("%Y-%m-%d") for d in date_cols_nf]
    bool_out = bool_matrix[date_cols_nf].copy(); bool_out.columns = date_col_names
    pct_out  = pct_matrix[date_cols_nf].copy();  pct_out.columns  = date_col_names
    today_col = date_col_names[0]

    # ── Assemble result ───────────────────────────────────────────────
    result = (
        pd.DataFrame({"Trending Days": trending})
        .join(chg_12d).join(chg_5d)
        .join(adx).join(rsi).join(rsi_short)
        .join(high_52w).join(pct_from_high)
        .join(near_high).join(rank_52w)
        .join(rs_score).join(weighted_rpi)
        .join(rpi_periods)
        .join(momentum)
        .join(macd).join(ema_df).join(ema_short)
    )
    result.index.name = "Symbol"
    result = result.reset_index()

    # Master joins FIRST (inner → filters ETFs)
    if not nse_master.empty:
        before = len(result)
        result = result.join(nse_master, on="Symbol", how="inner")
        if "Company Name" not in result.columns:
            result["Company Name"] = result["Symbol"]
        if "ISIN" not in result.columns:
            result["ISIN"] = ""
        logger.info("[%s] Master join: %d → %d (ETFs excluded)",
                    trade_date, before, len(result))
    else:
        result["Company Name"] = result["Symbol"]
        result["ISIN"] = ""

    if not sec_master.empty:
        result = result.join(sec_master, on="Symbol", how="left")
        if "Sector" in result.columns:
            result["Sector"] = result["Sector"].map(normalize_sector_name)

    for col in ["Company Name","ISIN","Sector","CapCategory"]:
        if col not in result.columns:
            result[col] = ""
        else:
            result[col] = result[col].fillna("").astype(str).str.strip()

    # Price columns (after joins so lengths match)
    result["Close"] = close_piv.iloc[:, -1].reindex(result["Symbol"].values).values
    result["High"]  = high_piv.iloc[:, -1].reindex(result["Symbol"].values).values
    result["Low"]   = low_piv.iloc[:, -1].reindex(result["Symbol"].values).values
    pct_today       = pct_out[today_col].reindex(result["Symbol"].values)
    result["chg_1d"] = pct_today.values

    # Bool matrix for Excel
    bool_out_r = bool_out.reindex(result["Symbol"].values)
    bool_out_r.index = result["Symbol"].values
    for dc in date_col_names:
        if dc in bool_out_r.columns:
            result[dc] = bool_out_r[dc].values

    # Volume
    today_raw = hist_to_date[hist_to_date["DATE"] == target_ts].set_index("SYMBOL")
    has_vol   = "TOTTRDQTY" in today_raw.columns

    logger.info("[%s] Final: %d symbols", trade_date, len(result))

    # ── Build Excel bytes (before upsert, needs full result + matrices) ─
    excel_bytes = None
    if return_excel:
        logger.info("[%s] Building Excel...", trade_date)
        excel_bytes = build_today_excel(
            result, bool_out_r, pct_out, date_col_names, trade_date
        )
        logger.info("[%s] Excel built: %d bytes", trade_date, len(excel_bytes))

    # ── DB WRITES ─────────────────────────────────────────────────────
    async with pool.acquire() as conn:

        # 1. Symbols
        sym_rows = [
            (str(r["Symbol"]),
             str(r.get("Company Name", r["Symbol"])),
             str(r.get("ISIN", "")),
             str(r.get("Sector", "")),
             str(r.get("CapCategory", "")))
            for _, r in result.iterrows()
        ]
        await conn.executemany(
            """
            INSERT INTO symbols (symbol, company_name, isin, sector, cap_category)
            VALUES ($1,$2,$3,$4,$5)
            ON CONFLICT (symbol) DO UPDATE SET
                company_name = excluded.company_name,
                isin         = excluded.isin,
                sector       = excluded.sector,
                cap_category = excluded.cap_category,
                is_active    = true,
                updated_at   = now()
            """,
            sym_rows,
        )

        # 2. trend_results — NO bool_matrix/pct_matrix stored
        trend_rows = []
        for _, row in result.iterrows():
            sym = row["Symbol"]
            vol    = _si(today_raw.at[sym, "TOTTRDQTY"])   if (has_vol and sym in today_raw.index) else None
            trades = _si(today_raw.at[sym, "TOTALTRADES"]) if (has_vol and sym in today_raw.index and "TOTALTRADES" in today_raw.columns) else None
            # Build bool_matrix and pct_matrix JSON
            bm = {}
            pm = {}
            for dc in date_col_names:
                if sym in bool_out.index:
                    bv = bool_out.at[sym, dc]
                    bm[dc] = bool(bv) if pd.notna(bv) else None
                if sym in pct_out.index:
                    pv = pct_out.at[sym, dc]
                    pm[dc] = round(float(pv), 4) if pd.notna(pv) else None
            trend_rows.append((
                trade_date, sym,                                        # $1  $2
                _si(row.get("Trending Days")),                          # $3
                _sf(row.get("chg_1d")),                                 # $4
                _sf(row.get("5d Change%")),                             # $5
                _sf(row.get("12d Change%")),                            # $6
                _sf(row.get("52W High")),                               # $7
                _sf(row.get("52W High %")),                             # $8
                _sb(row.get("Near 52W High")),                          # $9
                _sf(row.get("52W Rank")),                               # $10
                _sf(row.get("RSI 14")),                                 # $11
                _sf(row.get("RSI 1D")),                                 # $12
                _sf(row.get("RSI 1W")),                                 # $13
                _sf(row.get("ADX 14")),                                 # $14
                _sf(row.get("MACD Line")),                              # $15
                _sf(row.get("MACD Signal")),                            # $16
                _sf(row.get("MACD Hist")),                              # $17
                _sf(row.get("EMA 50")),                                 # $18
                _sf(row.get("EMA 200")),                                # $19
                str(row.get("EMA Signal") or ""),                       # $20
                _sf(row.get("EMA 9")),                                  # $21
                _sf(row.get("EMA 21")),                                 # $22
                _clamp(row.get("RS Score"),    1, 99),                  # $23
                _clamp(row.get("Weighted RPI"),1, 99),                  # $24
                _sf(row.get("RPI 2W")),                                 # $25
                _sf(row.get("RPI 3M")),                                 # $26
                _sf(row.get("RPI 6M")),                                 # $27
                _sf(row.get("RPI 6M SMA2W")),                           # $28
                _clamp(row.get("Momentum Score"), 0, 100),              # $29
                json.dumps(bm),                                         # $30 bool_matrix
                json.dumps(pm),                                         # $31 pct_matrix
                _sf(row.get("Close")),                                  # $32 open proxy
                _sf(row.get("High")),                                   # $33
                _sf(row.get("Low")),                                    # $34
                _sf(row.get("Close")),                                  # $35
                vol,                                                    # $36
                trades,                                                 # $37
            ))

        for start in range(0, len(trend_rows), _BATCH):
            await conn.executemany(
                """
                INSERT INTO trend_results (
                    trade_date, symbol,
                    trending_days,
                    chg_1d, chg_5d, chg_12d,
                    high_52w, pct_from_high, near_52w_high, rank_52w,
                    rsi_14, rsi_1d, rsi_1w, adx_14,
                    macd_line, macd_signal, macd_hist,
                    ema_50, ema_200, ema_signal,
                    ema_9, ema_21,
                    rs_score, weighted_rpi,
                    rpi_2w, rpi_3m, rpi_6m, rpi_6m_sma2w,
                    momentum_score,
                    bool_matrix, pct_matrix,
                    open_price, high_price, low_price, close_price,
                    volume, total_trades
                ) VALUES (
                    $1,$2,$3,
                    $4,$5,$6,$7,$8,$9,$10,
                    $11,$12,$13,$14,
                    $15,$16,$17,$18,$19,$20,
                    $21,$22,
                    $23,$24,$25,$26,$27,$28,
                    $29,
                    $30::jsonb,$31::jsonb,
                    $32,$33,$34,$35,$36,$37
                )
                ON CONFLICT (trade_date, symbol) DO UPDATE SET
                    trending_days  = excluded.trending_days,
                    chg_1d         = excluded.chg_1d,
                    chg_5d         = excluded.chg_5d,
                    chg_12d        = excluded.chg_12d,
                    high_52w       = excluded.high_52w,
                    pct_from_high  = excluded.pct_from_high,
                    near_52w_high  = excluded.near_52w_high,
                    rank_52w       = excluded.rank_52w,
                    rsi_14         = excluded.rsi_14,
                    rsi_1d         = excluded.rsi_1d,
                    rsi_1w         = excluded.rsi_1w,
                    adx_14         = excluded.adx_14,
                    macd_line      = excluded.macd_line,
                    macd_signal    = excluded.macd_signal,
                    macd_hist      = excluded.macd_hist,
                    ema_50         = excluded.ema_50,
                    ema_200        = excluded.ema_200,
                    ema_signal     = excluded.ema_signal,
                    ema_9          = excluded.ema_9,
                    ema_21         = excluded.ema_21,
                    rs_score       = excluded.rs_score,
                    weighted_rpi   = excluded.weighted_rpi,
                    rpi_2w         = excluded.rpi_2w,
                    rpi_3m         = excluded.rpi_3m,
                    rpi_6m         = excluded.rpi_6m,
                    rpi_6m_sma2w   = excluded.rpi_6m_sma2w,
                    momentum_score = excluded.momentum_score,
                    bool_matrix    = excluded.bool_matrix,
                    pct_matrix     = excluded.pct_matrix,
                    open_price     = excluded.open_price,
                    high_price     = excluded.high_price,
                    low_price      = excluded.low_price,
                    close_price    = excluded.close_price,
                    volume         = excluded.volume,
                    total_trades   = excluded.total_trades
                """,
                trend_rows[start : start + _BATCH],
            )
        logger.info("[%s] trend_results: %d rows", trade_date, len(trend_rows))

        # 3. sector_daily
        if "Sector" in result.columns:
            result_s = result.copy()
            result_s[today_col] = bool_out_r[today_col].values if today_col in bool_out_r.columns else False
            pct_s = pd.Series(pct_today.values, index=result["Symbol"].values)
            sec_12d   = build_sector_12d(result_s)
            sec_5d    = build_sector_5d(result_s)
            sec_today = build_sector_today(result_s, today_col, pct_s)
            sec = (
                sec_12d
                .merge(sec_5d[["Sector","Avg 5d Change%"]], on="Sector", how="outer")
                .merge(sec_today[["Sector","Stocks Up","Stocks Down","% Stocks Up","Avg Today Change%"]],
                       on="Sector", how="outer")
            )
            near_counts = (
                result_s[result_s.get("Near 52W High", pd.Series(False)) == True]
                .groupby("Sector")["Symbol"].count()
                if "Near 52W High" in result_s.columns else pd.Series(dtype=int)
            )
            sector_rows = []
            for _, r in sec.iterrows():
                sn = str(r.get("Sector","")).strip()
                if not sn: continue
                sc = int(r.get("Stocks",0) or 0)
                nh = int(near_counts.get(sn, 0))
                sector_rows.append((
                    trade_date, sn, sc,
                    int(r.get("Stocks Up",0) or 0),
                    int(r.get("Stocks Down",0) or 0),
                    _sf(r.get("% Stocks Up")),
                    _sf(r.get("Avg Trending Days")),
                    _sf(r.get("Avg 12d Change%")),
                    _sf(r.get("Avg 5d Change%")),
                    _sf(r.get("Avg Today Change%")),
                    _sf(r.get("Avg RSI 14")),
                    _sf(r.get("Avg ADX 14")),
                    _sf(r.get("Avg RS Score")),
                    _sf(r.get("Avg Momentum")),
                    nh,
                    round(nh/sc*100,1) if sc else None,
                ))
            await conn.executemany(
                """
                INSERT INTO sector_daily (
                    trade_date, sector,
                    stock_count, stocks_up, stocks_down, pct_stocks_up,
                    avg_trending_days, avg_chg_12d, avg_chg_5d, avg_chg_today,
                    avg_rsi_14, avg_adx_14, avg_rs_score, avg_momentum,
                    stocks_near_high, pct_near_high
                ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16)
                ON CONFLICT (trade_date, sector) DO UPDATE SET
                    stock_count       = excluded.stock_count,
                    stocks_up         = excluded.stocks_up,
                    stocks_down       = excluded.stocks_down,
                    pct_stocks_up     = excluded.pct_stocks_up,
                    avg_trending_days = excluded.avg_trending_days,
                    avg_chg_12d       = excluded.avg_chg_12d,
                    avg_chg_5d        = excluded.avg_chg_5d,
                    avg_chg_today     = excluded.avg_chg_today,
                    avg_rsi_14        = excluded.avg_rsi_14,
                    avg_adx_14        = excluded.avg_adx_14,
                    avg_rs_score      = excluded.avg_rs_score,
                    avg_momentum      = excluded.avg_momentum,
                    stocks_near_high  = excluded.stocks_near_high,
                    pct_near_high     = excluded.pct_near_high
                """,
                sector_rows,
            )

        # 4. market_calendar
        elapsed = round(time.monotonic() - t0, 2)
        await conn.execute(
            """
            INSERT INTO market_calendar
                (trade_date, is_trading_day, bhav_downloaded, engine_status,
                 symbol_count, engine_duration_secs, processed_at)
            VALUES ($1, true, true, 'done', $2, $3, now())
            ON CONFLICT (trade_date) DO UPDATE SET
                is_trading_day       = true,
                bhav_downloaded      = true,
                engine_status        = 'done',
                symbol_count         = excluded.symbol_count,
                engine_duration_secs = excluded.engine_duration_secs,
                processed_at         = now(),
                error_message        = null
            """,
            trade_date, len(result), elapsed,
        )

    logger.info("[%s] ✓ %d symbols  %.1fs", trade_date, len(result), elapsed)
    return {
        "symbols":     len(result),
        "elapsed":     elapsed,
        "trade_date":  str(trade_date),
        "excel_bytes": excel_bytes,
    }


# ════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════
async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date",      help="Trade date YYYY-MM-DD (default: today)")
    parser.add_argument("--recompute", action="store_true",
                        help="Wipe trend_results/sector_daily/market_calendar then recompute")
    args = parser.parse_args()

    db_url = os.environ.get("DATABASE_URL","")
    if not db_url:
        logger.error("DATABASE_URL not set"); sys.exit(1)

    trade_date = (datetime.date.fromisoformat(args.date)
                  if args.date else datetime.date.today())

    logger.info("="*55)
    logger.info("  TrendPulse — Compute Today  |  %s", trade_date)
    logger.info("="*55)

    pool = await asyncpg.create_pool(
        db_url, min_size=1, max_size=5,
        command_timeout=180,
        max_inactive_connection_lifetime=60,
    )

    # ── Optional full reset ────────────────────────────────────────────
    if args.recompute:
        logger.info("--recompute: wiping trend_results, sector_daily, market_calendar...")
        async with pool.acquire() as conn:
            await conn.execute("TRUNCATE trend_results")
            await conn.execute("TRUNCATE sector_daily")
            await conn.execute("TRUNCATE market_calendar")
            await conn.execute("UPDATE symbols SET is_active = true")
        logger.info("Tables cleared ✓")
    
    # ── Bhav cleanup (keep latest 255 dates only) ──────────────────────  ← ADD HERE
    async with pool.acquire() as conn:                                    #
        await cleanup_old_bhav_dates(conn)
    
    # ── Load data ──────────────────────────────────────────────────────
    async with pool.acquire() as conn:
        hist      = await load_price_history(conn, trade_date)
        nse_m, sec_m = await load_master_from_db(conn)

    # ── Compute ────────────────────────────────────────────────────────
    summary = await compute_and_upsert_today(
        pool, hist, nse_m, sec_m, trade_date
    )

    logger.info("="*55)
    logger.info("  DONE — %d symbols in %.1fs", summary["symbols"], summary["elapsed"])
    logger.info("="*55)
    await pool.close()


if __name__ == "__main__":
    asyncio.run(main())